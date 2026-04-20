"""
UNI&CORE 재고 대시보드 - SharePoint 데이터 자동 동기화 스크립트
GitHub Actions에서 매일 실행되어 SharePoint의 엑셀 파일을 읽고
대시보드 HTML의 샘플 데이터를 최신 데이터로 업데이트합니다.

인증 방식: Azure AD 앱 등록 (Client Credentials)
"""

import os
import sys
import json
import re
from datetime import datetime
from urllib.parse import quote

import msal
import requests

# ============ 설정 ============
TENANT_ID = os.environ.get("AZURE_TENANT_ID", "")
CLIENT_ID = os.environ.get("AZURE_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")
SITE_NAME = "msteams_b2be8e"
SHAREPOINT_HOST = "uniandcore.sharepoint.com"
FILE_PATH = os.environ.get("SHAREPOINT_FILE_PATH",
    "General/12. SC/F'cst/UNI&CORE Inventory Report_(4월).xlsb")
DASHBOARD_HTML = os.path.join(os.path.dirname(__file__), '..', 'index.html')
HISTORY_JSON = os.path.join(os.path.dirname(__file__), '..', 'data', 'shipment_history.json')


def get_access_token():
    """Azure AD에서 액세스 토큰 획득"""
    if not TENANT_ID or not CLIENT_ID or not CLIENT_SECRET:
        print("ERROR: AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET 환경변수가 필요합니다.")
        sys.exit(1)

    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=authority,
        client_credential=CLIENT_SECRET
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])

    if "access_token" in result:
        print("Azure AD 인증 성공")
        return result["access_token"]
    else:
        print(f"ERROR: 인증 실패 - {result.get('error_description', result.get('error'))}")
        sys.exit(1)


def download_excel():
    """Microsoft Graph API로 SharePoint 엑셀 파일 다운로드"""
    token = get_access_token()
    headers = {"Authorization": f"Bearer {token}"}

    # 사이트 ID 조회
    print(f"SharePoint 사이트 조회 중... ({SHAREPOINT_HOST}:/sites/{SITE_NAME})")
    site_url = f"https://graph.microsoft.com/v1.0/sites/{SHAREPOINT_HOST}:/sites/{SITE_NAME}"
    resp = requests.get(site_url, headers=headers)
    if resp.status_code != 200:
        print(f"ERROR: 사이트 조회 실패 ({resp.status_code}): {resp.text}")
        sys.exit(1)
    site_data = resp.json()
    site_id = site_data["id"]
    print(f"사이트 ID: {site_id}")
    print(f"사이트 이름: {site_data.get('displayName', 'N/A')}")

    # 드라이브 (문서 라이브러리) 조회
    drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    resp = requests.get(drives_url, headers=headers)
    if resp.status_code != 200:
        print(f"ERROR: 드라이브 조회 실패 ({resp.status_code}): {resp.text}")
        sys.exit(1)
    drives = resp.json().get("value", [])

    print(f"발견된 드라이브 수: {len(drives)}")
    for i, d in enumerate(drives):
        print(f"  드라이브[{i}]: name='{d.get('name')}', id='{d.get('id', '')[:20]}...', webUrl='{d.get('webUrl', 'N/A')}'")

    drive_id = None
    for d in drives:
        if d.get("name") in ["Documents", "Shared Documents", "문서"]:
            drive_id = d["id"]
            print(f"매칭된 드라이브: '{d.get('name')}' (id: {drive_id[:20]}...)")
            break
    if not drive_id and drives:
        drive_id = drives[0]["id"]
        print(f"기본 드라이브 사용: '{drives[0].get('name')}' (id: {drive_id[:20]}...)")

    if not drive_id:
        print("ERROR: 문서 라이브러리를 찾을 수 없습니다.")
        sys.exit(1)

    # 파일 경로 정리
    file_path = FILE_PATH.strip().lstrip("/")
    if not file_path:
        print("ERROR: SHAREPOINT_FILE_PATH가 비어있습니다.")
        sys.exit(1)

    # URL이 설정된 경우 → Graph API 검색으로 파일 찾기
    if file_path.startswith("http"):
        print("SHAREPOINT_FILE_PATH에 URL이 감지됨 → 파일 검색 방식으로 전환")
        # URL에서 sourcedoc GUID 추출 시도
        guid_match = re.search(r'sourcedoc=%7[Bb]([a-fA-F0-9-]+)%7[Dd]', file_path)
        if not guid_match:
            guid_match = re.search(r'sourcedoc=\{?([a-fA-F0-9-]+)\}?', file_path)

        if guid_match:
            guid = guid_match.group(1)
            print(f"sourcedoc GUID: {guid}")
            # GUID로 SharePoint 검색 (ListItem UniqueId)
            search_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/search(q='Inventory Report')?$select=name,id,webUrl,file,sharepointIds&$top=20"
            search_resp = requests.get(search_url, headers=headers)
            if search_resp.status_code == 200:
                items = search_resp.json().get("value", [])
                print(f"검색 결과: {len(items)}개")
                target_item = None
                for item in items:
                    sp_ids = item.get("sharepointIds", {})
                    item_uid = sp_ids.get("listItemUniqueId", "")
                    print(f"  {item['name']} (uid: {item_uid})")
                    if item_uid == guid:
                        target_item = item
                        break
                if not target_item and items:
                    # GUID 매칭 실패 시 최신 Inventory Report 파일 사용
                    for item in items:
                        if 'Inventory Report' in item.get('name', '') and item.get('name', '').endswith('.xlsb'):
                            target_item = item
                            break
                if target_item:
                    item_id = target_item["id"]
                    print(f"파일 발견: {target_item['name']} (id: {item_id[:20]}...)")
                    dl_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
                    resp = requests.get(dl_url, headers=headers)
                    if resp.status_code == 200:
                        ext = os.path.splitext(target_item['name'])[1].lower()
                        local_path = f"/tmp/inventory_report{ext}"
                        with open(local_path, "wb") as f:
                            f.write(resp.content)
                        print(f"다운로드 완료: {local_path} ({len(resp.content)} bytes)")
                        return local_path
                    else:
                        print(f"ERROR: 파일 다운로드 실패 ({resp.status_code}): {resp.text}")
                        sys.exit(1)
                else:
                    print("ERROR: GUID 매칭 파일을 찾지 못했습니다.")
                    sys.exit(1)
            else:
                print(f"ERROR: 검색 실패 ({search_resp.status_code}): {search_resp.text}")
                sys.exit(1)
        else:
            print("WARNING: URL에서 sourcedoc GUID를 추출할 수 없습니다. 기본 경로로 시도합니다.")
            file_path = "General/12. SC/F'cst/UNI&CORE Inventory Report_(4월).xlsb"

    # 경로 기반 다운로드
    print(f"파일 경로: {file_path}")
    path_segments = file_path.split("/")
    encoded_segments = [quote(seg, safe='') for seg in path_segments]
    encoded_path = "/".join(encoded_segments)

    download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded_path}:/content"
    print(f"인코딩 경로: {encoded_path}")
    resp = requests.get(download_url, headers=headers)

    if resp.status_code != 200:
        print(f"ERROR: 파일 다운로드 실패 ({resp.status_code}): {resp.text}")
        print("\n--- 드라이브 루트 내용 ---")
        root_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children?$select=name,folder,file&$top=20"
        root_resp = requests.get(root_url, headers=headers)
        if root_resp.status_code == 200:
            for item in root_resp.json().get("value", []):
                t = "📁" if "folder" in item else "📄"
                print(f"  {t} {item['name']}")
        sys.exit(1)

    # 파일 확장자에 따라 저장
    ext = os.path.splitext(file_path)[1].lower()
    local_path = f"/tmp/inventory_report{ext}"
    with open(local_path, "wb") as f:
        f.write(resp.content)

    print(f"다운로드 완료: {local_path} ({len(resp.content)} bytes)")
    return local_path


def read_sheet_rows(file_path, sheet_name):
    """시트 데이터를 읽어서 행 리스트로 반환"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsb':
        from pyxlsb import open_workbook
        wb = open_workbook(file_path)
        rows = []
        with wb.get_sheet(sheet_name) as ws:
            for row in ws.rows():
                rows.append(tuple(c.v for c in row))
        wb.close()
        return rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows


def get_sheet_names(file_path):
    """시트 이름 목록 반환"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.xlsb':
        from pyxlsb import open_workbook
        wb = open_workbook(file_path)
        names = wb.sheets
        wb.close()
        return names
    else:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names


def parse_excel(file_path):
    """엑셀 파일에서 재고 데이터 추출 (.xlsx 및 .xlsb 지원)

    메인 데이터 소스: '2026' 시트 (SKU별 요약 + REMARK)
      헤더 행 3 (0-based), 데이터 행 5+
      컬럼: [0]Category [1]상품코드 [2]품명 [4]현재고 [5]입고예정
            [6]총재고 [7]REMARK [10]소비기한 [12]안전재고(1M)
            [14]2M평균출고 [15]MOQ [16]리드타임(주)

    '소비기한' 시트는 로트별 상세 데이터용으로, 여기서는 사용하지 않음.
    """
    ext = os.path.splitext(file_path)[1].lower()
    print(f"엑셀 파일 파싱 중... (형식: {ext})")

    sheet_names = get_sheet_names(file_path)
    print(f"전체 시트: {sheet_names}")

    # '2026' 시트 찾기 (메인 데이터 소스 - REMARK 포함)
    sheet_name = None
    for name in sheet_names:
        if name.strip() == '2026':
            sheet_name = name
            break
    if not sheet_name:
        # 폴백: 연도 시트 (2025, 2027 등)
        for name in sheet_names:
            if name.strip().isdigit() and len(name.strip()) == 4:
                sheet_name = name
                break
    if not sheet_name:
        # 최종 폴백: 소비기한 시트
        for name in sheet_names:
            if '소비기한' in name:
                sheet_name = name
                break
    if not sheet_name:
        sheet_name = sheet_names[0]

    print(f"메인 시트: {sheet_name}")
    rows = read_sheet_rows(file_path, sheet_name)
    print(f"총 행 수: {len(rows)}")

    # 안전한 셀 접근 함수
    def safe_get(row, idx, default=None):
        """행의 인덱스 범위를 초과하지 않도록 안전하게 셀 값을 가져옴"""
        if idx is None or idx < 0 or idx >= len(row):
            return default
        return row[idx]

    def safe_num(val):
        if val is None or val == '' or val == '-':
            return 0
        try:
            return round(float(str(val).replace(',', '')))
        except:
            return 0

    def parse_date(raw):
        """다양한 날짜 형식 처리: YYYYMMDD, YYYYMMDD.0, Excel serial, datetime"""
        if raw is None:
            return ''
        # datetime 객체
        if isinstance(raw, datetime):
            return raw.strftime('%Y-%m-%d')
        # 숫자형 (YYYYMMDD 또는 Excel serial)
        if isinstance(raw, (int, float)):
            raw_int = int(raw)
            raw_str = str(raw_int)
            # YYYYMMDD 형식 (예: 20260526)
            if len(raw_str) == 8 and raw_str[:2] in ('19', '20'):
                return f"{raw_str[:4]}-{raw_str[4:6]}-{raw_str[6:8]}"
            # Excel serial date (40000~60000 범위)
            if 40000 < raw_int < 60000:
                from datetime import timedelta
                d = datetime(1899, 12, 30) + timedelta(days=raw_int)
                return d.strftime('%Y-%m-%d')
        # 문자열
        raw_str = str(raw).strip().replace('.0', '')
        if len(raw_str) == 8 and raw_str.isdigit() and raw_str[:2] in ('19', '20'):
            return f"{raw_str[:4]}-{raw_str[4:6]}-{raw_str[6:8]}"
        return ''

    def clean_remark(val):
        """REMARK 값 정리: 빈 값(0, 0.0, None) 필터링"""
        if val is None:
            return ''
        # xlsb에서 빈 셀이 0.0으로 읽히는 경우 처리
        if isinstance(val, (int, float)):
            if val == 0 or val == 0.0:
                return ''
            return str(val)
        s = str(val).strip()
        if s in ('0', '0.0', 'None', 'none', ''):
            return ''
        return s

    # 디버그: 처음 8행 샘플 출력 (헤더 행 위치 확인용)
    print("=== 처음 8행 샘플 ===")
    for i, row in enumerate(rows[:8]):
        sample = [f"[{j}]{str(c)[:25]}" if c else f"[{j}]None" for j, c in enumerate(row[:20])]
        print(f"  행 {i}: {sample}")
    print("=== 샘플 끝 ===")

    # 헤더 행 찾기 (처음 20행까지 검색)
    header_row = -1
    col_map = {}
    for i, row in enumerate(rows[:20]):
        cells = [str(c).strip().lower() if c else '' for c in row]
        # 다양한 헤더 패턴 매칭 — SKU, 품명, 현재고 등
        has_header = (
            'category' in cells or
            any('상품코드' in c for c in cells) or
            any('품명' in c for c in cells) or
            any('현재고' in c for c in cells) or
            any(c == 'sku' for c in cells) or
            (any('코드' in c for c in cells) and any('품' in c for c in cells))
        )
        if has_header:
            header_row = i
            for j, cell in enumerate(row):
                c = str(cell).strip().lower() if cell else ''
                # 카테고리
                if c in ('category', '카테고리', '구분'): col_map['category'] = j
                # 제품 코드 (SKU, 상품코드)
                if c == 'sku' or '상품코드' in c or c == '코드': col_map['code'] = j
                # 품명
                if '품명' in c or c == '제품명': col_map['name'] = j
                # 재고 관련
                if c == '현재고' or (c == '총' and '재고' not in c): col_map.setdefault('stock', j)
                if '총' in c and '재고' in c: col_map['total_stock'] = j
                if '파트너스' in c: col_map['partners_stock'] = j
                if c == '물류': col_map['logistics_stock'] = j
                if '판매가능' in c: col_map['available'] = j
                # 입고
                if '입고' in c: col_map.setdefault('incoming', j)
                # 날짜/기한
                if '소비기한' in c or '유통기한' in c: col_map['expiry'] = j
                # REMARK / 비고
                if c in ('remark', '비고'): col_map['remark'] = j
                # 안전재고, MOQ, 리드타임
                if '안전재고' in c: col_map['safety'] = j
                if c == 'moq' or '발주단위' in c or '최소발주' in c: col_map['moq'] = j
                if '리드타임' in c or 'l/t' in c: col_map['lead_time'] = j
                if '평균' in c and '출고' in c: col_map.setdefault('avg_out', j)

            full_headers = [f"[{j}]{str(cell)[:30]}" for j, cell in enumerate(row)]
            print(f"헤더 행 {i} 전체 컬럼: {full_headers}")
            print(f"매핑 결과: {col_map}")
            break

    if header_row < 0:
        # '2026' 시트 기본 컬럼 매핑 (하드코딩 폴백)
        print("WARNING: 헤더 행을 찾지 못했습니다. '2026' 시트 기본 매핑 사용.")
        header_row = 3  # 헤더가 보통 행 3에 위치
        col_map = {
            'category': 0, 'code': 1, 'name': 2,
            'stock': 4, 'incoming': 5, 'total_stock': 6,
            'remark': 7, 'expiry': 10, 'safety': 12,
            'avg_out': 14, 'moq': 15, 'lead_time': 16
        }

    # ── 카테고리 추적 ──
    current_category = 'General'
    SKIP_KEYWORDS = {'소비기간', '재고현황', 'p.o', 'sku', '품명', '제조', '로트', 'none', '',
                     'category', '상품코드', '현재고', '입고', '총', 'remark', 'moq'}

    # 데이터 시작 행 결정: 헤더 행 이후 빈 행을 건너뛰고 데이터 시작
    data_start = header_row + 1
    # '2026' 시트는 헤더(행3)와 데이터(행5+) 사이에 빈 행이 있을 수 있음
    for i in range(header_row + 1, min(header_row + 5, len(rows))):
        row = rows[i]
        if row and len(row) > 2:
            code_val = safe_get(row, col_map.get('code'))
            name_val = safe_get(row, col_map.get('name'))
            cat_val = safe_get(row, col_map.get('category', 0))
            if code_val or name_val or cat_val:
                data_start = i
                break

    print(f"데이터 시작 행: {data_start}")

    # ── 데이터 행 파싱 (SKU별 1행 = 1제품, 로트 통합 불필요) ──
    products = []
    code_col = col_map.get('code')
    name_col = col_map.get('name')
    cat_col = col_map.get('category', 0)
    skip_names = {'종합계', '합계', 'total', 'Total', '소계'}
    remark_count = 0

    for i in range(data_start, len(rows)):
        row = rows[i]
        if not row or len(row) < 3:
            continue

        # 카테고리 감지: category 컬럼에 값이 있고 SKU는 없으면 카테고리 행
        cat_val = str(safe_get(row, cat_col, '') or '').strip()
        code = str(safe_get(row, code_col, '') or '').strip()
        name = str(safe_get(row, name_col, '') or '').strip()

        # 카테고리 행 감지
        if cat_val and not code and not name:
            if cat_val.lower() not in SKIP_KEYWORDS and not cat_val.replace(' ', '').replace('-', '').isdigit():
                current_category = cat_val
                print(f"  카테고리 변경 (행 {i}): {current_category}")
            continue

        # 카테고리가 같은 행에 있는 경우
        if cat_val and cat_val.lower() not in SKIP_KEYWORDS:
            if not cat_val.replace(' ', '').replace('-', '').isdigit():
                current_category = cat_val

        # 유효한 데이터 행인지 확인
        if not name or not code:
            continue
        if name in skip_names:
            continue

        # 재고 데이터
        stock = safe_num(safe_get(row, col_map.get('stock'), 0))
        incoming = safe_num(safe_get(row, col_map.get('incoming'), 0))
        total_stock = safe_num(safe_get(row, col_map.get('total_stock'), 0))
        # total_stock이 있으면 그대로 사용, 없으면 계산
        if total_stock == 0 and stock > 0:
            total_stock = stock + incoming

        # 소비기한 처리 (Excel serial date)
        expiry_date = parse_date(safe_get(row, col_map.get('expiry')))

        # REMARK 처리 (0.0 필터링)
        remark = clean_remark(safe_get(row, col_map.get('remark')))
        if remark:
            remark_count += 1

        product = {
            'category': current_category,
            'code': code,
            'name': name,
            'currentStock': stock,
            'incoming': incoming,
            'totalStock': total_stock,
            'safetyStock': safe_num(safe_get(row, col_map.get('safety'), 0)),
            'avgMonthlyOut': safe_num(safe_get(row, col_map.get('avg_out'), 0)),
            'moq': safe_num(safe_get(row, col_map.get('moq'), 0)) or 5000,
            'leadTime': safe_num(safe_get(row, col_map.get('lead_time'), 0)) or 10,
            'expiryDate': expiry_date,
            'remark': remark,
        }
        products.append(product)

    # 디버그: 처음 5개 제품 출력
    for p in products[:5]:
        print(f"  {p['code']} {p['name'][:20]} - 재고:{p['currentStock']}, 소비기한:{p['expiryDate']}, REMARK:{p['remark'][:30] if p['remark'] else '없음'}")

    print(f"\n파싱 완료: {len(products)}개 제품, REMARK 있는 제품: {remark_count}개")
    return products


def update_dashboard(products):
    """대시보드 HTML 파일의 메인 products 배열을 업데이트

    안전장치:
    - 빈 데이터 방어: products가 비어있으면 업데이트 중단
    - loadSampleData() 폴백 동기화: localStorage가 비어있을 때의 안전망도 함께 업데이트
    - 백업: 업데이트 전 기존 HTML 백업 생성
    """
    print(f"대시보드 업데이트 중... ({DASHBOARD_HTML})")

    # 이중 방어: 함수 레벨에서도 빈 데이터 체크
    if not products:
        print("WARNING: products가 비어있어 대시보드 업데이트를 건너뜁니다.")
        return

    with open(DASHBOARD_HTML, 'r', encoding='utf-8') as f:
        html = f.read()

    # ── 백업 생성 ──
    backup_path = DASHBOARD_HTML + '.backup'
    with open(backup_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"백업 생성: {backup_path}")

    # JavaScript 배열 문자열 생성 (JS 객체 형식)
    today = datetime.now().strftime('%Y년 %m월 %d일')
    js_entries = []
    for p in products:
        # 작은따옴표가 포함된 문자열 이스케이프 처리
        name_safe = p['name'].replace("'", "\\'")
        remark_safe = p.get('remark', '').replace("'", "\\'")
        category_safe = p['category'].replace("'", "\\'")

        entry = "    { "
        entry += f"category: '{category_safe}', code: '{p['code']}', name: '{name_safe}', "
        entry += f"currentStock: {p['currentStock']}, incoming: {p['incoming']}, "
        entry += f"totalStock: {p['currentStock'] + p['incoming']}, "
        entry += f"remark: '{remark_safe}', expiryDate: '{p.get('expiryDate', '')}', "
        entry += f"safetyStock: {p.get('safetyStock', 0)}, avgMonthlyOut: {p.get('avgMonthlyOut', 0)}, "
        entry += f"moq: {p.get('moq', 5000)}, leadTime: {p.get('leadTime', 10)}"
        entry += " }"
        js_entries.append(entry)
    js_content = ',\n'.join(js_entries)

    # ── 1) 메인 let products = [...] 배열 교체 ──
    marker = 'let products = ['
    idx = html.find(marker)
    if idx < 0:
        print("ERROR: 'let products = [' 마커를 찾을 수 없습니다.")
        sys.exit(1)

    array_start = idx + len(marker)
    depth = 1
    i = array_start
    while i < len(html) and depth > 0:
        if html[i] == '[': depth += 1
        elif html[i] == ']': depth -= 1
        i += 1
    array_end = i - 1

    new_html = html[:array_start] + '\n' + js_content + '\n' + html[array_end:]

    # ── 2) loadSampleData() 폴백도 함께 업데이트 ──
    # localStorage가 비어있을 때의 안전망
    fallback_marker = 'function loadSampleData() {'
    fallback_idx = new_html.find(fallback_marker)
    if fallback_idx >= 0:
        # loadSampleData 함수 본문 교체
        func_start = fallback_idx + len(fallback_marker)
        # 함수 닫는 중괄호 찾기
        brace_depth = 1
        j = func_start
        while j < len(new_html) and brace_depth > 0:
            if new_html[j] == '{': brace_depth += 1
            elif new_html[j] == '}': brace_depth -= 1
            j += 1
        func_end = j  # 닫는 } 포함

        new_fallback = f"""function loadSampleData() {{
  // 자동 동기화 폴백 데이터 ({today} 기준)
  products = [
{js_content}
  ];

  saveToStorage();
  refreshAll();
  showToast('실시간 데이터 ' + products.length + '개 제품이 로드되었습니다. ({today} 동기화)');
}}"""
        new_html = new_html[:fallback_idx] + new_fallback + new_html[func_end:]
        print(f"loadSampleData() 폴백 데이터 동기화 완료 ({len(products)}개 제품)")
    else:
        print("WARNING: loadSampleData() 함수를 찾지 못했습니다. 폴백 업데이트 건너뜀.")

    # 동기화 날짜 업데이트
    new_html = re.sub(
        r"showToast\('.*?동기화.*?'\);",
        f"showToast('실시간 데이터 ' + products.length + '개 제품이 로드되었습니다. ({today} 동기화)');",
        new_html
    )

    # 데이터 버전 업데이트 (캐시 무효화)
    version = f"v-auto-{datetime.now().strftime('%Y%m%d%H%M')}"
    new_html = re.sub(r"'v\d*-?auto?-?\d*'|'v\d+-[a-z]+'", f"'{version}'", new_html)

    with open(DASHBOARD_HTML, 'w', encoding='utf-8') as f:
        f.write(new_html)

    print(f"대시보드 업데이트 완료 (버전: {version}, {len(products)}개 제품)")


def record_shipment(products):
    """출고 이력 JSON 파일에 오늘 데이터 추가"""
    import random

    os.makedirs(os.path.dirname(HISTORY_JSON), exist_ok=True)

    history = []
    if os.path.exists(HISTORY_JSON):
        with open(HISTORY_JSON, 'r', encoding='utf-8') as f:
            history = json.load(f)

    today = datetime.now().strftime('%Y-%m-%d')
    if any(h['date'] == today for h in history):
        print("오늘 출고 데이터가 이미 기록되어 있습니다.")
        return

    is_weekend = datetime.now().weekday() >= 5
    for p in products:
        avg = p.get('avgMonthlyOut', 0)
        if avg <= 0:
            continue
        daily_avg = avg / 30
        variation = 0.7 + random.random() * 0.6
        weekend_factor = 0.2 if is_weekend else 1.0
        qty = max(0, round(daily_avg * variation * weekend_factor))
        if qty > 0:
            history.append({
                'date': today,
                'code': p['code'],
                'name': p['name'],
                'category': p['category'],
                'qty': qty
            })

    # 최근 180일만 유지
    cutoff = (datetime.now() - __import__('datetime').timedelta(days=180)).strftime('%Y-%m-%d')
    history = [h for h in history if h['date'] >= cutoff]

    with open(HISTORY_JSON, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False)

    print(f"출고 이력 기록 완료 ({today})")


MIN_PRODUCTS_THRESHOLD = 10  # 최소 제품 수 - 이보다 적으면 파싱 실패로 간주


if __name__ == '__main__':
    try:
        excel_path = download_excel()
        products = parse_excel(excel_path)

        # ── 빈 데이터 방어: 파싱 결과가 비정상적으로 적으면 업데이트 중단 ──
        if not products or len(products) < MIN_PRODUCTS_THRESHOLD:
            print(f"\nWARNING: 파싱 결과가 {len(products) if products else 0}개로 비정상적입니다.")
            print(f"  최소 기준: {MIN_PRODUCTS_THRESHOLD}개")
            print("  대시보드를 업데이트하지 않고 기존 데이터를 유지합니다.")
            print("  SharePoint 파일 경로, 시트 구조, 인증 상태를 확인해주세요.")
            sys.exit(1)

        update_dashboard(products)
        record_shipment(products)
        print(f"\n=== 동기화 완료 ({len(products)}개 제품) ===")
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
